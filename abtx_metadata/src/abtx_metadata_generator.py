import numpy as np
import pandas
import string
from dateutil import parser
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from qiimp import SAMPLETYPE_SHORTHAND_KEY, QC_NOTE_KEY, DO_NOT_USE_VAL, \
    LEAVE_BLANK_VAL, HOST_SUBJECT_ID_KEY, SAMPLE_NAME_KEY, \
    COLLECTION_TIMESTAMP_KEY, \
    extract_config_dict, deepcopy_dict, \
    merge_sample_and_subject_metadata, \
    write_extended_metadata_from_df

# internal code keys
PLATE_SAMPLE_ID_KEY = "plate_sample_id"
PLATE_ROW_ID_KEY = "plate_row_id"
PLATE_COL_ID_KEY = "plate_col_id"
PLATE_ID_KEY = "plate_id"
PLATING_NOTES_KEY = "plating_notes"
PLATING_DATE_KEY = "plating_date"
SUBJECT_SHORTHAND_KEY = "subject_shorthand"

# config keys
DESIRED_PLATES_KEY = "desired_plates"
CONSTANT_FIELDS_KEY = "constant_fields"
SUBJECT_SPECIFIC_METADATA_KEY = "subject_specific_metadata"
DOB_KEY = "date_of_birth"
STUDY_START_DATE_KEY = "study_start_date"
LOCATION_BREAK_KEY = "location_break"
BEFORE_LOCATION_KEY = "before_location"
AFTER_LOCATION_KEY = "after_location"
AFTER_START_DATE_KEY = "after_start_date"
LOCATIONS_KEY = "locations"
LOCATION_KEY = "location"
ASSUME_DATES_PRESENT_KEY = "assume_dates_present"

# metadata keys
COUNTRY_KEY = "country"
IS_COLLECTION_TIMESTAMP_VALID_KEY = "is_collection_timestamp_valid"
ORDINAL_TIMESTAMP_KEY = "ordinal_timestamp"
ORIGINAL_COLLECTION_TIMESTAMP_KEY = "original_collection_timestamp"
COLLECTION_DATE_KEY = "collection_date"
DESCRIPTION_KEY = "description"
HOST_AGE_KEY = "host_age"
M03_AGE_YEARS_KEY = "m03_age_years"
DAYS_SINCE_FIRST_DAY_KEY = "days_since_first_day"
NOTES_KEY = "notes"

# constant field values
BLANK_VAL = "blank"

# columns added to the metadata that are not actually part of it
INTERNAL_COL_KEYS = [PLATE_ID_KEY, PLATE_ROW_ID_KEY, PLATE_COL_ID_KEY,
                     PLATE_SAMPLE_ID_KEY, PLATING_NOTES_KEY, PLATING_DATE_KEY,
                     SUBJECT_SHORTHAND_KEY, SAMPLETYPE_SHORTHAND_KEY,
                     QC_NOTE_KEY]


def make_abtx_extendable_metadata_df(
        platemap_fp, included_sheet_names, subject_metadata_fp, config):

    found_platemaps_by_num = _get_platemap_dfs_by_num_from_file(
        platemap_fp, included_sheet_names, config)

    desired_plate_dfs = []
    for curr_platemap_id, curr_platemap_obj in found_platemaps_by_num.items():
        # TODO: remove debug
        print(curr_platemap_id)
        if curr_platemap_id == "Loding":
            print("DEBUG")

        # if we want all plates or we want certain plates including this one
        if DESIRED_PLATES_KEY not in config or \
                curr_platemap_id in config[DESIRED_PLATES_KEY]:
            curr_transformed_df = _transform_platemap_df(curr_platemap_obj)
            if curr_transformed_df is not None:
                desired_plate_dfs.append(curr_transformed_df)
        # endif this is a desired plate
    # next plate

    if DESIRED_PLATES_KEY in config:
        missing_plate_nums = [x for x in config[DESIRED_PLATES_KEY]
                              if x not in found_platemaps_by_num.keys()]
        if len(missing_plate_nums) > 0:
            raise ValueError(
                f"Desired plates not found in platemap file: "
                f"{sorted(missing_plate_nums)}")
        # end if there are missing plates
    # end if specific desired plates are specified

    total_desired_plates_df = pandas.concat(desired_plate_dfs, ignore_index=True)

    metadata_df = _mine_plate_sample_id_for_metadata(total_desired_plates_df)
    subject_metadata_df = pandas.read_csv(subject_metadata_fp)  #, dtype=str)

    merged_metadata_df = merge_sample_and_subject_metadata(
        metadata_df, subject_metadata_df, SUBJECT_SHORTHAND_KEY)

    merged_metadata_df = _set_metadata_for_blanks(merged_metadata_df)
    merged_metadata_df = _set_notes(merged_metadata_df)

    full_metadata_df = _update_metadata_by_subjects(merged_metadata_df, config)

    return full_metadata_df


class AnnotatedPlatemap:
    COL_NAMES = [x for x in range(1, 13)]
    ROW_NAMES = ["A", "B", "C", "D", "E", "F", "G", "H"]

    def __init__(self, fields_lines, metadata_fields, plate_num_str,
                 assume_dates_present):
        self.cell_df = None
        self.plater_initials = None
        self.plating_date = None

        if fields_lines is not None:
            self.cell_df = pandas.DataFrame(
                fields_lines,
                index=self.ROW_NAMES, columns=self.COL_NAMES)

        self._plate_id_field = metadata_fields.pop(0)
        if assume_dates_present:
            self._plating_initials_and_date_field = metadata_fields.pop(0)

            plating_info_txt = self._plating_initials_and_date_field.value
            if plating_info_txt:
                plating_info = self._plating_initials_and_date_field.value.split()
                self.plater_initials = plating_info[0]
                self.plating_date = plating_info[1]

        self.metadata_fields = metadata_fields
        self.plate_num_str = plate_num_str


def _get_platemap_dfs_by_num_from_file(path, included_sheet_names, config):
    workbook = load_workbook(filename=path)

    found_platemaps_by_num = {}

    curr_platemap_num_str = None
    curr_platemap_lines = None
    curr_platemap_metadatas = None

    blank_platemap_name = "()"
    assume_dates_present = config.get(ASSUME_DATES_PRESENT_KEY, False)

    def _finish_platemap(platemap_lines, metadata_fields, platemap_num_str):
        if curr_platemap_lines is not None:
            found_platemap = AnnotatedPlatemap(
                platemap_lines, metadata_fields, platemap_num_str,
                assume_dates_present)
            if platemap_num_str != blank_platemap_name:
                if platemap_num_str in found_platemaps_by_num:
                    raise ValueError(
                        f"Plate # {platemap_num_str} found more than once")
                # endif found plate >1x

                found_platemaps_by_num[platemap_num_str] = found_platemap
            # endif platemap_num_str is not the blank plate placeholder
        return None, None, None

    for sheet_name in workbook.sheetnames:
        if sheet_name not in included_sheet_names:
            continue

        sheet = workbook[sheet_name]

        # if there is one last platemap from the previous sheet that
        # hasn't been finished, finish it up
        curr_platemap_lines, curr_platemap_metadatas, curr_platemap_num_str = \
            _finish_platemap(curr_platemap_lines, curr_platemap_metadatas,
                             curr_platemap_num_str)

        for row in sheet.rows:
            curr_first_field = row[0].value
            if curr_first_field is None:
                curr_first_field = ""
            else:
                curr_first_field = curr_first_field.strip()

            # capture plate row info
            if curr_first_field in AnnotatedPlatemap.ROW_NAMES:
                if curr_first_field == "A":
                    if curr_platemap_lines is not None:
                        raise ValueError("Found beginning of a new platemap when old one is not finished")
                    curr_platemap_lines = []
                # endif curr first field is A

                # leave out column letter and only take *up to* (but not
                # including--split ends are exclusive) to field w index 13
                # (field index 12 is 12th col of platemap bc we're ignoring
                # 1st col but the indexing is 0-based, so it evens out)
                # because only taking 96-well plates, which are
                # 12 col by 8 rows
                curr_platemap_lines.append(row[1:13])

                # look for any metadata fields--i.e., any filled fields beyond
                # the end of the platemap--and collect any found
                curr_cell_index = 13
                curr_putative_metadata_cell = row[curr_cell_index]
                while curr_putative_metadata_cell.value is not None:
                    curr_platemap_metadatas.append(curr_putative_metadata_cell)
                    curr_cell_index += 1
                    curr_putative_metadata_cell = row[curr_cell_index]
            # throw error if >96-well plate
            elif curr_first_field == "I":
                raise ValueError(f"Only 96-well plate maps expected: {curr_first_field}; current platemap ignored")
            # collect existing platemap info if reached empty cell
            elif curr_first_field == "":
                curr_platemap_lines, curr_platemap_metadatas, curr_platemap_num_str = \
                    _finish_platemap(
                        curr_platemap_lines, curr_platemap_metadatas,
                        curr_platemap_num_str)
            # look for the plate name
            else:
                if curr_first_field.startswith("Plate"):
                    temp_str = curr_first_field.replace("#", "")
                    temp_str = temp_str.replace("Plate", "")
                    number_str_pieces = temp_str.strip().split()
                    if len(number_str_pieces) > 0:
                        number_str = number_str_pieces[0]
                        if number_str != "":
                            curr_platemap_num_str = number_str
                elif curr_platemap_lines is None:
                    field_pieces = curr_first_field.split("ABTX_")
                    if len(field_pieces) == 2:
                        curr_platemap_num_str = field_pieces[1]
                    else:
                        curr_platemap_num_str = curr_first_field
                else:
                    print(f"Ignoring first field: '{curr_first_field}'")

                if curr_platemap_num_str:
                    plate_name_cell = row[0]
                    curr_platemap_metadatas = [plate_name_cell]

                    # If 13th cell (0-based index 12) holds the integer 12,
                    # that means we are definitely looking at the top row of
                    # a plate definition (which lists the column numbers 1-12).
                    # Therefore, the 14th cell (0-based index 13) will be the
                    # one holding the plater's initials and the plating date.
                    if row[12].value == 12:
                        plating_date_cell = row[13]
                        curr_platemap_metadatas.append(plating_date_cell)
            # endif examine first cell in row
        # endif next row
    # endif next sheet name

    return found_platemaps_by_num


def _get_field_notes(a_field, notes_by_color):
    field_notes = None
    field_notes_list = []

    if a_field.comment:
        comment_text = a_field.comment.text.replace('\t', '').replace('\n', '')
        field_notes_list.append(comment_text)

    field_color_index = a_field.fill.bgColor.index
    color_notes = notes_by_color.get(field_color_index, [])
    field_notes_list.extend(color_notes)

    if len(field_notes_list) > 0:
        field_notes = ";".join(field_notes_list)
    return field_notes


def _transform_platemap_df(platemap_obj):
    # assume we start from a pandas dataframe of a platemap:
    # Plate#153 ORAL	1	2   etc
    # A	5.20.18.RK.T	5.17.18.RK.T    etc
    # B	5.19.18.RK.T	5.16.18.RK.T
    # C	5.13.18.RK.T	5.15.18.RK.T
    # D	5.18.18.RK.T	7.28.18.RK.T
    # E	5.14.18.RK.T	7.25.18.RK.T
    # F	7.15.18.RK.T	7.17.18.RK.T
    # G	7.22.18.RK.T	7.21.18.RK.T
    # H	7.12.18.RK.T	7.11.18.RK.T
    # let's assume column numbers (e.g. 1-12) are column "names" and
    # row letters (e.g. A-H) are indices

    # gather color-based metadata
    notes_by_color_index = {}
    for curr_metadata_field in platemap_obj.metadata_fields:
        curr_bg_color = curr_metadata_field.fill.bgColor.index
        if curr_bg_color != "00000000":
            curr_field_text = curr_metadata_field.value
            if curr_field_text is None:
                continue
            existing_color_notes = notes_by_color_index.get(curr_bg_color, [])
            existing_color_notes.append(curr_field_text)
            notes_by_color_index[curr_bg_color] = existing_color_notes
        # endif background color is not white
    # next field of metadata

    # for each column "name"
    dfs_to_concat = []
    platemap_df = platemap_obj.cell_df
    for curr_col_id in platemap_df:
        curr_col_fields = platemap_df.loc[:, curr_col_id]
        notes_series = curr_col_fields.apply(
            _get_field_notes, notes_by_color=notes_by_color_index)
        notes_series = notes_series.reset_index(drop=True)
        notes_series.fillna(LEAVE_BLANK_VAL, inplace=True)

        curr_col_vals = curr_col_fields.apply(lambda x: x.value)
        curr_col_df = pandas.DataFrame(curr_col_vals)
        curr_col_df.reset_index(inplace=True)
        curr_col_df.rename(
            columns={"index": PLATE_ROW_ID_KEY,
                     curr_col_id: PLATE_SAMPLE_ID_KEY},
            inplace=True)

        curr_col_df[PLATE_COL_ID_KEY] = curr_col_id
        curr_col_df[PLATING_NOTES_KEY] = notes_series
        curr_col_df[PLATING_DATE_KEY] = platemap_obj.plating_date
        dfs_to_concat.append(curr_col_df)
    # next column

    # concat all the dataframes for columns in the plate
    plate_df = pandas.concat(dfs_to_concat, ignore_index=True)
    plate_df[PLATE_ID_KEY] = platemap_obj.plate_num_str

    # check if all PLATE_SAMPLE_ID_KEY values in the dataframe are None
    empty_sample_ids = plate_df[PLATE_SAMPLE_ID_KEY].isna()
    all_sample_ids_empty = empty_sample_ids.all()
    if all_sample_ids_empty:
        plate_df = None

    return plate_df


# def _populate_metadata_df(plates_df, config):
#     metadata_df = plates_df.copy()
#     metadata_df[QC_NOTE_KEY] = LEAVE_BLANK_VAL
#
#     metadata_df = _update_metadata_from_config_dict(metadata_df, config)
#     metadata_df = _mine_plate_sample_id_for_metadata(metadata_df)
#     metadata_df = _set_metadata_for_blanks(metadata_df)
#     metadata_df = _set_notes(metadata_df)
#     metadata_df = _update_metadata_for_subject_and_sample_type(
#         metadata_df, config)
#     return metadata_df


def _mine_plate_sample_id_for_metadata(metadata_df):
    putative_sample_name = _construct_sample_name(metadata_df)
    metadata_df[SAMPLE_NAME_KEY] = putative_sample_name

    # Grab the "original collection timestamp" from the part of the sample
    # name that comes before the first letter
    before_first_letter_regex = r"(.*?)[A-z].*$"
    # strip trailing period from column contents
    original_collection_strs = putative_sample_name.str.extract(
            before_first_letter_regex, expand=False)
    original_collection_strs = \
        original_collection_strs.str.replace(r"\.$", "", regex=True)
    time_str = putative_sample_name.str.extract(r"(\d+)$", expand=False)
    original_collection_strs = \
        original_collection_strs + " " + time_str.fillna("")
    original_collection_strs = original_collection_strs.str.strip()
    original_collection_strs[original_collection_strs == ""] = np.NAN
    metadata_df[ORIGINAL_COLLECTION_TIMESTAMP_KEY] = original_collection_strs

    # Dig the subject and sample type shorthands out of the part of the sample
    # name that is at or after the first letter
    # e.g., in 5.20.18.RK.T, look at RK.T
    first_letter_and_after_regex = r".*?([A-z].*)$"
    text_info = putative_sample_name.str.extract(
        first_letter_and_after_regex, expand=False)
    text_info = text_info.str.lower()
    # replace any entry in the text_info series that starts with DO_NOT_USE_VAL
    # with just the DO_NOT_USE_VAL
    text_info = text_info.str.replace(
        r"^" + DO_NOT_USE_VAL + r".*$", DO_NOT_USE_VAL, regex=True)
    text_info_pieces = text_info.str.split(r"\.", expand=True)

    putative_subject_shorthands = text_info_pieces[0].str.strip().copy()
    putative_subject_shorthands.fillna(DO_NOT_USE_VAL, inplace=True)
    blanks_mask = putative_subject_shorthands.str.startswith(BLANK_VAL)
    putative_subject_shorthands[blanks_mask] = BLANK_VAL
    metadata_df[SUBJECT_SHORTHAND_KEY] = putative_subject_shorthands
    # metadata_df[SUBJECT_SHORTHAND_KEY] = \
    #     metadata_df[SUBJECT_SHORTHAND_KEY].fillna(DO_NOT_USE_VAL)

    metadata_df[SAMPLETYPE_SHORTHAND_KEY] = text_info_pieces[1].str.strip()
    metadata_df[SAMPLETYPE_SHORTHAND_KEY] = \
        metadata_df[SAMPLETYPE_SHORTHAND_KEY].str.rstrip(string.digits).copy()

    return metadata_df


def _set_metadata_for_blanks(metadata_df):
    blank_host_mask = metadata_df[SUBJECT_SHORTHAND_KEY] == BLANK_VAL
#        metadata_df[SUBJECT_SHORTHAND_KEY].str.startswith(BLANK_VAL)

    metadata_df.loc[blank_host_mask, HOST_SUBJECT_ID_KEY] = \
        metadata_df.loc[blank_host_mask, DESCRIPTION_KEY] = \
        metadata_df.loc[blank_host_mask, SAMPLE_NAME_KEY]
    #metadata_df.loc[blank_host_mask, SUBJECT_SHORTHAND_KEY] = BLANK_VAL
    metadata_df.loc[blank_host_mask, SAMPLETYPE_SHORTHAND_KEY] = BLANK_VAL

    # This provisionally sets the collection date to the plating date.
    # This date will be run through the usual date handling later on (to e.g.
    # check that it is after the study start date, fill in the ordinal
    # timestamp, etc).
    metadata_df.loc[blank_host_mask, COLLECTION_TIMESTAMP_KEY] = \
        metadata_df.loc[blank_host_mask, PLATING_DATE_KEY]

    return metadata_df


def _construct_sample_name(metadata_df):
    # scrub the plate_sample_id column
    putative_sample_names = metadata_df[PLATE_SAMPLE_ID_KEY].copy()
    putative_sample_names = putative_sample_names.str.strip()
    putative_sample_names = putative_sample_names.str.replace(" ", ".")
    putative_sample_names = putative_sample_names.str.replace(",", ".")
    # collapse multiple periods into one
    putative_sample_names = putative_sample_names.str.replace(
        r"\.+", ".", regex=True)

    # if there are any empties in putative_sample_names, fill them with
    # "donotuse" + plate num + "." + col num + row letter
    empty_names_mask = \
        (putative_sample_names == "") | (putative_sample_names.isna())
    if empty_names_mask.any():
        putative_sample_names[empty_names_mask] = \
            _add_plate_well_info_to_selected_names(
                metadata_df, empty_names_mask, DO_NOT_USE_VAL)

    # TODO: Need to find out how these blanks are represented in data already
    #  in Qiita (bc that can't be changed)!
    # TODO: Ask Gail if period betw BLANK and plate name is important here/
    #  important NOT to appear for donotuse ...
    # Find any blanks that aren't yet in the correct format
    blank_names_mask = putative_sample_names.str.lower() == BLANK_VAL
    if blank_names_mask.any():
        putative_sample_names[blank_names_mask] = \
            _add_plate_well_info_to_selected_names(
                metadata_df, blank_names_mask,
                putative_sample_names[blank_names_mask])

    # # TODO: this is a hack; should we keep this or do something else with
    # #  these problem names?
    # misnames = {"tA": "t.A", "TA": "T.A", "tB": "t.B", "TB": "T.B",
    #             "lhA": "lh.A", "LHA": "LH.A", "lhB": "lh.B", "LHB": "LH.B",
    #             "fhA": "fh.A", "FHA": "FH.A", "fhB": "fh.B", "FHB": "FH.B",
    #             "RHA": "RH.A", "RHB": "RH.B"}
    # for curr_key, curr_val in misnames.items():
    #     putative_sample_names = putative_sample_names.str.replace(
    #         curr_key, curr_val, regex=True)

    # keep = False means mark all duplicates as such, including first one
    duplicates_mask = putative_sample_names.duplicated(keep=False)
    if duplicates_mask.any():
        duplicate_names = putative_sample_names[duplicates_mask]
        duplicates_counts = duplicate_names.groupby(duplicate_names).cumcount()
        if max(duplicates_counts) > 26:
            raise ValueError(
                "Cannot disambiguate more than 26 identical names")
        duplicate_letters = duplicates_counts.apply(
            lambda x: string.ascii_uppercase[x])
        disambiguated_names = duplicate_names + "." + duplicate_letters
        putative_sample_names[duplicates_mask] = disambiguated_names

    return putative_sample_names


def _add_plate_well_info_to_selected_names(metadata_df, names_mask, name_base):
    result = name_base + \
        metadata_df.loc[names_mask, PLATE_ID_KEY].astype("string") + \
        "." + \
        metadata_df.loc[names_mask, PLATE_COL_ID_KEY].astype("string") \
        + metadata_df.loc[names_mask, PLATE_ROW_ID_KEY].astype("string")
    return result


def _set_notes(metadata_df):
    # In the future, may want to also include some other source in the notes
    # contents, which is why this gets its own function.  For now, though, just
    # put any plating notes into the official notes field
    metadata_df[NOTES_KEY] = metadata_df[PLATING_NOTES_KEY]
    return metadata_df


def _update_metadata_by_subjects(metadata_df, config):
    # gather global settings
    settings_dict = {STUDY_START_DATE_KEY: config.get(STUDY_START_DATE_KEY)}

    subject_dfs = []
    subject_shorthands = pandas.unique(metadata_df[SUBJECT_SHORTHAND_KEY])
    for curr_subject_shorthand in subject_shorthands:
        concatted_dfs = _generate_metadata_for_subject(
            metadata_df, curr_subject_shorthand, settings_dict, config)

        subject_dfs.append(concatted_dfs)
    # next subject

    output_df = pandas.concat(subject_dfs, ignore_index=True)
    output_df.replace(LEAVE_BLANK_VAL, "", inplace=True)
    return output_df


def _generate_metadata_for_subject(
        metadata_df, subject_shorthand, settings_dict, config):

    subject_mask = \
        metadata_df[SUBJECT_SHORTHAND_KEY] == subject_shorthand
    subject_df = metadata_df.loc[subject_mask, :].copy()

    known_host_shorthands = config[SUBJECT_SPECIFIC_METADATA_KEY].keys()
    if subject_shorthand not in known_host_shorthands:
        subject_df[QC_NOTE_KEY] = "invalid subject shorthand"
    else:
        # gather subject-level settings
        subject_specific_dict = \
            config[SUBJECT_SPECIFIC_METADATA_KEY][subject_shorthand]
        curr_settings_dict = deepcopy_dict(settings_dict)
        curr_settings_dict.update(subject_specific_dict)

        # # update the metadata with the constant values defined for this subject
        # subject_df = _update_metadata_from_config_consts(
        #     subject_df, curr_settings_dict)

        # add/validate date fields based on subject-specific settings
        study_start_str = curr_settings_dict[STUDY_START_DATE_KEY]
        subject_df = _add_collection_dates(subject_df, study_start_str)
        subject_df = _add_days_since_start(subject_df, study_start_str)
        if DOB_KEY in curr_settings_dict:
            subject_df = _add_host_age(
                subject_df, curr_settings_dict[DOB_KEY])

        # add fixed location metadata if this subject has a fixed location
        # (note that this doesn't HAVE to come after the date handling,
        # logically, but it makes sense to put it near the location_break
        # handling for multiple locations based on date, and *that* DOES have
        # to come after the date handling).
        if LOCATION_KEY in subject_specific_dict:
            subj_location_name = subject_specific_dict[LOCATION_KEY]
            subj_location_dict = config[LOCATIONS_KEY][subj_location_name]
            subject_df = _update_metadata_from_config_consts(
                subject_df, subj_location_dict)
        # endif this subject has a fixed location

        # add time-dependent location metadata if a location-date break exists
        if LOCATION_BREAK_KEY in curr_settings_dict:
            subject_df = _add_location_info_from_break(
                subject_df, curr_settings_dict[LOCATION_BREAK_KEY], config)
        # endif this subject has more than one location, based on date
    # endif subject shorthand is valid

    return subject_df


def _update_metadata_from_config_consts(metadata_df, config_section_dict):
    constant_fields_dict = config_section_dict.get(CONSTANT_FIELDS_KEY)
    if constant_fields_dict:
        output_df = metadata_df.copy()
        for curr_field_name, curr_const_val in constant_fields_dict.items():
            output_df[curr_field_name] = curr_const_val
        metadata_df = output_df
    return metadata_df


# def _generate_metadata_for_sample_type(
#         subject_df, sample_type_shorthand, curr_settings_dict,
#         sample_types_specific_dict, config):
#
#     # get df of records for this sample type
#     sample_type_mask = \
#         subject_df[SAMPLETYPE_SHORTHAND_KEY] == sample_type_shorthand
#     sample_type_df = subject_df.loc[sample_type_mask, :].copy()
#
#     known_sample_type_shorthands = sample_types_specific_dict.keys()
#     if sample_type_shorthand not in known_sample_type_shorthands:
#         sample_type_df[QC_NOTE_KEY] = "invalid sample type shorthand"
#     else:
#         # get sample-type-specific metadata dict
#         sample_type_specific_dict = \
#             sample_types_specific_dict[sample_type_shorthand]
#         sample_type_alias = sample_type_specific_dict.get(ALIAS_KEY)
#         if sample_type_alias:
#             sample_type_specific_dict = \
#                 sample_types_specific_dict[sample_type_alias]
#
#         sample_type_df = _update_metadata_from_config_dict(
#             sample_type_df, sample_type_specific_dict)
#
#
#
#     return sample_type_df


def _add_collection_dates(plates_df, study_start_date_str):
    output_df = plates_df.copy()
    collection_dates = _get_collection_dates(plates_df, study_start_date_str)
    output_df[COLLECTION_TIMESTAMP_KEY] = collection_dates
    output_df[COLLECTION_DATE_KEY] = collection_dates
    output_df[ORDINAL_TIMESTAMP_KEY] = collection_dates.dt.strftime('%Y%m%d')
    none_date_mask = output_df[COLLECTION_TIMESTAMP_KEY].isna()
    output_df[IS_COLLECTION_TIMESTAMP_VALID_KEY] = ~none_date_mask
    output_df.loc[none_date_mask, QC_NOTE_KEY] = \
        "invalid/unparseable date"

    return output_df


def _get_collection_dates(plates_df, start_date_str):
    start_date = parser.parse(start_date_str, dayfirst=False)

    def _get_date_from_sample_id_if_empty(row):
        sample_date = None
        putative_date_str = None

        existing_date_str = row[COLLECTION_TIMESTAMP_KEY]
        if existing_date_str and pandas.notna(existing_date_str):
            putative_date_str = existing_date_str
        else:
            sample_id = row[SAMPLE_NAME_KEY]
            fname_pieces = sample_id.split(".")
            if len(fname_pieces) >= 3:
                putative_date_str = "/".join(fname_pieces[:3])

        if putative_date_str:
            try:
                sample_date = parser.parse(
                    putative_date_str, fuzzy=True, dayfirst=False)
            except:
                pass

        # sanity check: can't be before study start date
        if sample_date and sample_date < start_date:
            sample_date = None

        return sample_date

    collection_dates = \
        plates_df.apply(_get_date_from_sample_id_if_empty, axis=1)
    return collection_dates


def _add_days_since_start(plates_df, study_start_str):
    output_df = plates_df.copy()
    days_since = _get_duration_since_date(
        plates_df, study_start_str, return_years=False)
    output_df[DAYS_SINCE_FIRST_DAY_KEY] = days_since
    return output_df


def _get_duration_since_date(plates_df, start_date_str, return_years=True):
    start_date = parser.parse(start_date_str, dayfirst=False)

    def _get_diff_in_years(x):
        result = None
        input_date = pandas.to_datetime(x)
        try:
            if return_years:
                a_relativedelta = relativedelta(input_date, start_date)
                result = a_relativedelta.years
            else:
                a_timedelta = input_date - start_date
                result = a_timedelta.days
            result = int(result)
        except:
            pass
        return result

    age_in_years = plates_df[COLLECTION_TIMESTAMP_KEY].apply(_get_diff_in_years)
    return age_in_years


def _add_host_age(plates_df, dob_str):
    output_df = plates_df.copy()
    age_in_years = _get_duration_since_date(plates_df, dob_str)
    output_df[HOST_AGE_KEY] = age_in_years
    output_df[M03_AGE_YEARS_KEY] = age_in_years
    return output_df


def _add_location_info_from_break(plates_df, location_break_dict, config):
    output_df = plates_df.copy()
    before_location_name = location_break_dict[BEFORE_LOCATION_KEY]
    after_location_name = location_break_dict[AFTER_LOCATION_KEY]
    first_after_date_str = location_break_dict[AFTER_START_DATE_KEY]
    first_after_date = parser.parse(first_after_date_str, dayfirst=False)

    before_location_fields = \
        config[LOCATIONS_KEY][before_location_name][CONSTANT_FIELDS_KEY]
    before_mask = plates_df[COLLECTION_TIMESTAMP_KEY] < first_after_date
    for a_key, a_val in before_location_fields.items():
        output_df.loc[before_mask, a_key] = a_val

    after_location_fields = \
        config[LOCATIONS_KEY][after_location_name][CONSTANT_FIELDS_KEY]
    for a_key, a_val in after_location_fields.items():
        output_df.loc[~before_mask, a_key] = a_val
    return output_df


if __name__ == "__main__":
    # TODO: remove hardcoded arguments
    a_platemap_fp = "/Users/abirmingham/Downloads/Rob_ABTX_Updated.xlsx"
    included_sheet_names_list = ["Platemaps"]
    subject_metadata_fp = "/Users/abirmingham/Work/Repositories/custom_abtx_metadata_generator/abtx_subject_metadata.csv"
    output_dir = "/Users/abirmingham/Desktop/"
    output_base = "scraped_ABTX_metadata"

    config_dict = extract_config_dict(None, starting_fp=__file__)
    extendable_metadata_df = make_abtx_extendable_metadata_df(
        a_platemap_fp, included_sheet_names_list, subject_metadata_fp,
        config_dict)

    # # TODO: remove write for qiimp debugging
    # extendable_metadata_df.to_csv(
    #     "/Users/abirmingham/Desktop/extended_abtx_metadata.csv",
    #     index=False)

    write_extended_metadata_from_df(
        extendable_metadata_df, config_dict, output_dir, output_base,
        study_specific_transformers_dict=None, suppress_empty_fails=True)
